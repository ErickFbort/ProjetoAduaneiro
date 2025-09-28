"""
API REST para veículos
"""

from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required
from app.models.veiculo import Veiculo
from app import db
import csv
import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

veiculos_bp = Blueprint('veiculos', __name__)

@veiculos_bp.route('/veiculos', methods=['GET', 'POST'])
@login_required
def api_veiculos():
    """API para listar e criar veículos"""
    if request.method == 'POST':
        data = request.get_json()
        
        # Verificar se placa já existe
        if Veiculo.query.filter_by(placa=data['placa']).first():
            return jsonify({'error': 'Placa já cadastrada'}), 400
        
        veiculo = Veiculo(
            motorista_responsavel=data['motorista_responsavel'],
            cpf_motorista=data['cpf_motorista'],
            placa=data['placa'],
            renavam=data.get('renavam'),
            tipo=data['tipo'],
            tipo_outros=data.get('tipo_outros'),
            estado=data.get('estado'),
            municipio=data.get('municipio'),
            observacoes=data.get('observacoes'),
            status=data.get('status', 'active')
        )
        
        db.session.add(veiculo)
        db.session.commit()
        
        return jsonify({'message': 'Veículo criado com sucesso'}), 201
    
    elif request.method == 'GET':
        veiculos = Veiculo.query.all()
        return jsonify([veiculo.to_dict() for veiculo in veiculos])

@veiculos_bp.route('/veiculos/<int:veiculo_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_veiculo_detail(veiculo_id):
    """API para operações específicas de veículo"""
    veiculo = Veiculo.query.get_or_404(veiculo_id)
    
    if request.method == 'GET':
        return jsonify(veiculo.to_dict())
    
    elif request.method == 'PUT':
        data = request.get_json()
        
        # Verificar se placa já existe (exceto para o próprio veículo)
        existing_veiculo = Veiculo.query.filter_by(placa=data['placa']).first()
        if existing_veiculo and existing_veiculo.id != veiculo_id:
            return jsonify({'error': 'Placa já cadastrada'}), 400
        
        veiculo.motorista_responsavel = data['motorista_responsavel']
        veiculo.cpf_motorista = data['cpf_motorista']
        veiculo.placa = data['placa']
        veiculo.renavam = data.get('renavam')
        veiculo.tipo = data['tipo']
        veiculo.tipo_outros = data.get('tipo_outros')
        veiculo.estado = data.get('estado')
        veiculo.municipio = data.get('municipio')
        veiculo.observacoes = data.get('observacoes')
        veiculo.status = data.get('status', veiculo.status)
        
        db.session.commit()
        return jsonify({'message': 'Veículo atualizado com sucesso'})
    
    elif request.method == 'DELETE':
        db.session.delete(veiculo)
        db.session.commit()
        return jsonify({'message': 'Veículo excluído com sucesso'})

@veiculos_bp.route('/veiculos/export', methods=['GET'])
@login_required
def export_veiculos():
    """Exportar veículos para Excel formatado"""
    try:
        # Buscar todos os veículos
        veiculos = Veiculo.query.all()
        
        # Criar workbook e worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Veículos"
        
        # Definir cabeçalhos
        headers = [
            'ID', 'Motorista Responsável', 'CPF Motorista', 'Placa', 'RENAVAM',
            'Tipo', 'Tipo Outros', 'Estado', 'Município', 'Observações', 'Status', 'Data Criação'
        ]
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Adicionar dados
        for row, veiculo in enumerate(veiculos, 2):
            data = [
                veiculo.id,
                veiculo.motorista_responsavel,
                veiculo.cpf_motorista,
                veiculo.placa,
                veiculo.renavam or '',
                veiculo.tipo,
                veiculo.tipo_outros or '',
                veiculo.estado or '',
                veiculo.municipio or '',
                veiculo.observacoes or '',
                veiculo.status,
                veiculo.created_at.strftime('%d/%m/%Y %H:%M:%S') if veiculo.created_at else ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Alternar cor de fundo das linhas
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Ajustar largura das colunas
        column_widths = [8, 25, 18, 12, 15, 15, 15, 8, 20, 30, 12, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Congelar primeira linha (cabeçalho)
        ws.freeze_panes = "A2"
        
        # Criar arquivo em memória
        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'veiculos_export_{timestamp}.xlsx'
        
        return send_file(
            mem,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao exportar veículos: {str(e)}'}), 500

@veiculos_bp.route('/veiculos/import', methods=['POST'])
@login_required
def import_veiculos():
    """Importar veículos de CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not file.filename.lower().endswith('.csv'):
            return jsonify({'error': 'Arquivo deve ser CSV'}), 400
        
        # Ler arquivo CSV
        csv_data = file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_data))
        
        imported_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, start=2):  # Começa em 2 porque linha 1 é cabeçalho
            try:
                # Validar dados obrigatórios
                if not row.get('Motorista Responsável') or not row.get('CPF Motorista') or not row.get('Placa') or not row.get('Tipo'):
                    errors.append(f'Linha {row_num}: Campos obrigatórios não preenchidos')
                    continue
                
                # Verificar se placa já existe
                if Veiculo.query.filter_by(placa=row['Placa']).first():
                    errors.append(f'Linha {row_num}: Placa {row["Placa"]} já cadastrada')
                    continue
                
                # Criar veículo
                veiculo = Veiculo(
                    motorista_responsavel=row['Motorista Responsável'],
                    cpf_motorista=row['CPF Motorista'],
                    placa=row['Placa'],
                    renavam=row.get('RENAVAM', '').strip() or None,
                    tipo=row['Tipo'],
                    tipo_outros=row.get('Tipo Outros', '').strip() or None,
                    estado=row.get('Estado', '').strip() or None,
                    municipio=row.get('Município', '').strip() or None,
                    observacoes=row.get('Observações', '').strip() or None,
                    status=row.get('Status', 'active')
                )
                
                db.session.add(veiculo)
                imported_count += 1
                
            except Exception as e:
                errors.append(f'Linha {row_num}: {str(e)}')
        
        # Salvar no banco
        db.session.commit()
        
        return jsonify({
            'message': f'Importação concluída! {imported_count} veículos importados com sucesso.',
            'imported_count': imported_count,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro ao importar veículos: {str(e)}'}), 500
