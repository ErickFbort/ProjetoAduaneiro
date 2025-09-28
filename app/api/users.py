"""
API REST para usuários
"""

from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required
from app.models.user import User
from app import db
import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

users_bp = Blueprint('users', __name__)

@users_bp.route('/users', methods=['GET', 'POST'])
@login_required
def api_users():
    """API para listar e criar usuários"""
    if request.method == 'POST':
        data = request.get_json()
        
        # Verificar se email já existe
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email já cadastrado'}), 400
        
        # Verificar se CPF já existe
        if User.query.filter_by(cpf=data['cpf']).first():
            return jsonify({'error': 'CPF já cadastrado'}), 400
        
        user = User(
            name=data['name'],
            lastname=data['lastname'],
            cpf=data['cpf'],
            email=data['email'],
            status=data.get('status', 'active'),
            group=data['group'],
            permissions=data.get('permissions', '[]')
        )
        user.set_password(data['password'])
        
        db.session.add(user)
        db.session.commit()
        
        return jsonify({'message': 'Usuário criado com sucesso'}), 201
    
    elif request.method == 'GET':
        users = User.query.all()
        return jsonify([user.to_dict() for user in users])

@users_bp.route('/users/<int:user_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_user_detail(user_id):
    """API para operações específicas de usuário"""
    user = User.query.get_or_404(user_id)
    
    if request.method == 'GET':
        return jsonify(user.to_dict())
    
    elif request.method == 'PUT':
        data = request.get_json()
        
        # Verificar se email já existe (exceto para o próprio usuário)
        existing_user = User.query.filter_by(email=data['email']).first()
        if existing_user and existing_user.id != user_id:
            return jsonify({'error': 'Email já cadastrado'}), 400
        
        # Verificar se CPF já existe (exceto para o próprio usuário)
        existing_user = User.query.filter_by(cpf=data['cpf']).first()
        if existing_user and existing_user.id != user_id:
            return jsonify({'error': 'CPF já cadastrado'}), 400
        
        user.name = data['name']
        user.lastname = data['lastname']
        user.cpf = data['cpf']
        user.email = data['email']
        user.status = data.get('status', user.status)
        user.group = data['group']
        user.permissions = data.get('permissions', '[]')
        
        if 'password' in data and data['password']:
            user.set_password(data['password'])
        
        db.session.commit()
        return jsonify({'message': 'Usuário atualizado com sucesso'})
    
    elif request.method == 'DELETE':
        db.session.delete(user)
        db.session.commit()
        return jsonify({'message': 'Usuário removido com sucesso'})

@users_bp.route('/users/export', methods=['GET'])
@login_required
def export_users():
    """Exportar usuários para Excel"""
    try:
        users = User.query.all()
        
        # Criar DataFrame
        data = []
        for user in users:
            data.append({
                'Nome': user.name,
                'Sobrenome': user.lastname,
                'CPF': user.cpf,
                'E-mail': user.email,
                'Grupo': user.group,
                'Status': 'Ativo' if user.status == 'active' else 'Bloqueado'
            })
        
        df = pd.DataFrame(data)
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuários"
        
        # Adicionar dados
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Ajustar largura das colunas
        column_widths = {'A': 20, 'B': 20, 'C': 15, 'D': 30, 'E': 20, 'F': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Salvar em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='usuarios.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao exportar usuários: {str(e)}'}), 500

@users_bp.route('/users/import', methods=['POST'])
@login_required
def import_users():
    """Importar usuários de arquivo CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({'error': 'Apenas arquivos CSV são permitidos'}), 400
        
        # Ler arquivo CSV
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_input = csv.DictReader(stream)
        
        imported_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_input, start=2):
            try:
                # Validar campos obrigatórios
                required_fields = ['Nome', 'Sobrenome', 'CPF', 'E-mail', 'Grupo']
                for field in required_fields:
                    if not row.get(field):
                        errors.append(f'Linha {row_num}: Campo "{field}" é obrigatório')
                        continue
                
                # Verificar se usuário já existe
                if User.query.filter_by(email=row['E-mail']).first():
                    errors.append(f'Linha {row_num}: E-mail já cadastrado')
                    continue
                
                if User.query.filter_by(cpf=row['CPF']).first():
                    errors.append(f'Linha {row_num}: CPF já cadastrado')
                    continue
                
                # Criar usuário
                user = User(
                    name=row['Nome'],
                    lastname=row['Sobrenome'],
                    cpf=row['CPF'],
                    email=row['E-mail'],
                    group=row['Grupo'],
                    status=row.get('Status', 'Ativo').lower() == 'ativo' and 'active' or 'blocked',
                    permissions='[]'
                )
                user.set_password('123456')  # Senha padrão
                
                db.session.add(user)
                imported_count += 1
                
            except Exception as e:
                errors.append(f'Linha {row_num}: {str(e)}')
        
        if imported_count > 0:
            db.session.commit()
        
        return jsonify({
            'message': f'{imported_count} usuários importados com sucesso',
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao importar usuários: {str(e)}'}), 500
