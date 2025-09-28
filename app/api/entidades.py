"""
API REST para entidades
"""

from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required
from app.models.entidade import Entidade
from app import db
import csv
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

entidades_bp = Blueprint('entidades', __name__)

@entidades_bp.route('/entidades', methods=['GET', 'POST'])
@login_required
def api_entidades():
    """API para listar e criar entidades"""
    if request.method == 'POST':
        data = request.get_json()
        
        # Verificar se CPF/CNPJ já existe
        if Entidade.query.filter_by(cpf_cnpj=data['cpf_cnpj']).first():
            return jsonify({'error': 'CPF/CNPJ já cadastrado'}), 400
        
        entidade = Entidade(
            razao_social=data['razao_social'],
            cpf_cnpj=data['cpf_cnpj'],
            nome_fantasia=data['nome_fantasia'],
            inscricao_estadual=data.get('inscricao_estadual'),
            tipo_cliente=data['tipo_cliente'],
            pagamento=data['pagamento'],
            retencao=data.get('retencao', False),
            valor_retencao=data.get('valor_retencao'),
            prazo_retencao=data.get('prazo_retencao'),
            relatorio_nd=data.get('relatorio_nd', False),
            email_faturamento=data['email_faturamento'],
            email_operacional=data['email_operacional'],
            email_despachante=data['email_despachante'],
            telefone=data.get('telefone'),
            notificacoes=data.get('notificacoes', '[]'),
            tipos=data.get('tipos', '[]'),
            status=data.get('status', 'active')
        )
        
        db.session.add(entidade)
        db.session.commit()
        
        return jsonify({'message': 'Entidade criada com sucesso'}), 201
    
    elif request.method == 'GET':
        entidades = Entidade.query.all()
        return jsonify([entidade.to_dict() for entidade in entidades])

@entidades_bp.route('/entidades/<int:entidade_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_entidade_detail(entidade_id):
    """API para operações específicas de entidade"""
    entidade = Entidade.query.get_or_404(entidade_id)
    
    if request.method == 'GET':
        return jsonify(entidade.to_dict())
    
    elif request.method == 'PUT':
        data = request.get_json()
        
        # Verificar se CPF/CNPJ já existe (exceto para a própria entidade)
        existing_entidade = Entidade.query.filter_by(cpf_cnpj=data['cpf_cnpj']).first()
        if existing_entidade and existing_entidade.id != entidade_id:
            return jsonify({'error': 'CPF/CNPJ já cadastrado'}), 400
        
        entidade.razao_social = data['razao_social']
        entidade.cpf_cnpj = data['cpf_cnpj']
        entidade.nome_fantasia = data['nome_fantasia']
        entidade.inscricao_estadual = data.get('inscricao_estadual')
        entidade.tipo_cliente = data['tipo_cliente']
        entidade.pagamento = data['pagamento']
        entidade.retencao = data.get('retencao', False)
        entidade.valor_retencao = data.get('valor_retencao')
        entidade.prazo_retencao = data.get('prazo_retencao')
        entidade.relatorio_nd = data.get('relatorio_nd', False)
        entidade.email_faturamento = data['email_faturamento']
        entidade.email_operacional = data['email_operacional']
        entidade.email_despachante = data['email_despachante']
        entidade.telefone = data.get('telefone')
        entidade.notificacoes = data.get('notificacoes', '[]')
        entidade.tipos = data.get('tipos', '[]')
        entidade.status = data.get('status', entidade.status)
        
        db.session.commit()
        return jsonify({'message': 'Entidade atualizada com sucesso'})
    
    elif request.method == 'DELETE':
        db.session.delete(entidade)
        db.session.commit()
        return jsonify({'message': 'Entidade excluída com sucesso'})

@entidades_bp.route('/entidades/export', methods=['GET'])
@login_required
def export_entidades():
    """Exportar entidades para Excel"""
    try:
        entidades = Entidade.query.all()
        
        # Criar DataFrame
        data = []
        for entidade in entidades:
            pagamento_text = {
                '001': 'Vista',
                '002': 'Correntista', 
                '003': 'Express'
            }.get(entidade.pagamento, entidade.pagamento)
            
            data.append({
                'Razão Social': entidade.razao_social,
                'Nome Fantasia': entidade.nome_fantasia,
                'CPF/CNPJ': entidade.cpf_cnpj,
                'Tipo Cliente': entidade.tipo_cliente,
                'Inscrição Estadual': entidade.inscricao_estadual or '',
                'Pagamento': pagamento_text,
                'Telefone': entidade.telefone or '',
                'Email Faturamento': entidade.email_faturamento,
                'Email Operacional': entidade.email_operacional,
                'Email Despachante': entidade.email_despachante,
                'Retenção': 'Sim' if entidade.retencao else 'Não',
                'Valor Retenção': entidade.valor_retencao or '',
                'Prazo Retenção': entidade.prazo_retencao or '',
                'Relatório ND': 'Sim' if entidade.relatorio_nd else 'Não',
                'Status': 'Ativo' if entidade.status == 'active' else 'Bloqueado'
            })
        
        df = pd.DataFrame(data)
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Entidades"
        
        # Adicionar dados
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adicionar bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Ajustar largura das colunas
        column_widths = {
            'A': 25, 'B': 25, 'C': 18, 'D': 15, 'E': 18, 'F': 12,
            'G': 15, 'H': 25, 'I': 25, 'J': 25, 'K': 10, 'L': 15,
            'M': 15, 'N': 12, 'O': 10
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='entidades.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao exportar entidades: {str(e)}'}), 500

@entidades_bp.route('/entidades/import', methods=['POST'])
@login_required
def import_entidades():
    """Importar entidades do CSV"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({'error': 'Arquivo deve ser CSV'}), 400
        
        # Ler CSV
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_input = csv.DictReader(stream)
        
        entidades_criadas = 0
        erros = []
        
        for row_num, row in enumerate(csv_input, start=2):
            try:
                # Validar campos obrigatórios
                if not row.get('Razão Social') or not row.get('CPF/CNPJ'):
                    erros.append(f"Linha {row_num}: Razão Social e CPF/CNPJ são obrigatórios")
                    continue
                
                # Verificar se CPF/CNPJ já existe
                if Entidade.query.filter_by(cpf_cnpj=row['CPF/CNPJ']).first():
                    erros.append(f"Linha {row_num}: CPF/CNPJ {row['CPF/CNPJ']} já cadastrado")
                    continue
                
                # Mapear pagamento
                pagamento_map = {
                    'Vista': '001',
                    'Correntista': '002', 
                    'Express': '003'
                }
                pagamento = pagamento_map.get(row.get('Pagamento', ''), '001')
                
                # Criar entidade
                entidade = Entidade(
                    razao_social=row['Razão Social'],
                    nome_fantasia=row.get('Nome Fantasia', ''),
                    cpf_cnpj=row['CPF/CNPJ'],
                    tipo_cliente=row.get('Tipo Cliente', 'Pessoa Fisica'),
                    inscricao_estadual=row.get('Inscrição Estadual', ''),
                    pagamento=pagamento,
                    telefone=row.get('Telefone', ''),
                    email_faturamento=row.get('Email Faturamento', ''),
                    email_operacional=row.get('Email Operacional', ''),
                    email_despachante=row.get('Email Despachante', ''),
                    retencao=row.get('Retenção', '').lower() == 'sim',
                    valor_retencao=float(row.get('Valor Retenção', 0)) if row.get('Valor Retenção') else None,
                    prazo_retencao=int(row.get('Prazo Retenção', 0)) if row.get('Prazo Retenção') else None,
                    relatorio_nd=row.get('Relatório ND', '').lower() == 'sim',
                    status='active' if row.get('Status', '').lower() == 'ativo' else 'blocked',
                    notificacoes='[]',
                    tipos='[]'
                )
                
                db.session.add(entidade)
                entidades_criadas += 1
                
            except Exception as e:
                erros.append(f"Linha {row_num}: {str(e)}")
        
        db.session.commit()
        
        message = f"Importação concluída! {entidades_criadas} entidades criadas."
        if erros:
            message += f" {len(erros)} erros encontrados."
        
        return jsonify({
            'message': message,
            'entidades_criadas': entidades_criadas,
            'errors': erros
        })
        
    except Exception as e:
        return jsonify({'error': f'Erro ao importar entidades: {str(e)}'}), 500
